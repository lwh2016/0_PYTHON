# -*- coding: utf-8 -*-

# from openpyxl import Workbook
# import datetime
# import time

# wb = Workbook()  # 创建文件对象

# ws = wb.active  # 获取第一个sheet

# # Data can be assigned directly to cells
# ws['A1'] = 42  # 写入数字
# ws['B1'] = "你好" + "automation test"  # 写入中文（unicode中文也可）

# # Rows can also be appended
# ws.append([1, 2, 3])  # 写入多个单元格

# # Python types will automatically be converted

# ws['A2'] = datetime.datetime.now()  #写入一个当前时间
# # 写入一个自定义的时间格式
# ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())

# # Save the file
# wb.save("f:\\sample.xlsx")
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('f:\\TestSummaryReport.xlsx')
ws = wb.active
rows = []
for row in ws.iter_rows():
    rows.append(row)

print(rows)
# print(len(rows))  # 所有行
# # print(rows[0])  # 获取第一行
# print(rows[3][0])  # 获取第一行第一列的单元格对象
# print(rows[3][1].value)  # 获取第一行第一列的单元格对象的值
# print(len(rows[39]))

# print(rows[len(rows) - 1])  # 获取最后行 print rows[-1]
# print(rows[len(rows) - 1][len(rows[0]) - 1])  # 获取最后一行和最后一列的单元格对象
# print(rows[len(rows) - 1][len(rows[0]) - 1].value)  # 获取第后一行和最后一列的单元格对象的值
