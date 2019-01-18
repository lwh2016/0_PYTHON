import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook


class Title(object):
    def __init__(self, level, content, style):
        self.level = level
        self.content = content
        self.style = style


class Table(object):
    def __init__(self, title, content, style):
        self.title = title
        self.content = content
        self.style = style


class chart(object):
    def __init__(self, index, content, style):
        pass


class Report(object):
    def __init__(self):
        pass


class ReportForm(object):
    def __init__(self, templateFile):
        self.template = load_workbook(templateFile).active
        self.rows = []
        for row in self.template.iter_rows():
            self.rows.append(row)

    def getTemplateTable(self, templateTable):
        for row in self.rows:
            if row[0].value == templateTable:
                print(row[0].value)


class GenSumRept(object):
    def __init__(self, report, name):
        self.report = report
        self.xlsxFile = xlsxwriter.Workbook(name)
        self.sheet = self.xlsxFile.add_worksheet(name)

    def write2Excel(self):
        pass


def main():
    reptForm = ReportForm(
        'F:\\0_PYTHON\\3-PythonModule\\2-Excel\TestSummaryReport.xlsx')
    tableTileList = [
        '1. Test basic information', '2. Test result summary',
        {
            '3. Test data':
            ['3.1. Test items', '1.3. Test summary and statistics data']
        }, '4. Test issues'
    ]
    reptForm.getTemplateTable()


if __name__ == '__main__':
    main()